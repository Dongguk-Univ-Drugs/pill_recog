import numpy as np
import cv2
import sys
import openpyxl as op

redLow = (0, 60, 80)
redHigh = (45, 255, 255)
greenLow = (45, 15, 10)
greenHigh = (80, 255, 255)
blueLow = (90, 60, 70)
blueHigh = (115, 255, 255)
blackLow = (0, 0, 0)
blackHigh = (180, 255, 40)
whiteLow = (0, 0, 220)
whiteHigh = (180, 40, 255)


def get_color(src, low, high, color, color_arr):
    src_hsv = cv2.cvtColor(src, cv2.COLOR_BGR2HSV)

    src_hsv = cv2.inRange(src_hsv, low, high)

    # 필요하면 주석 제거
    # kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (8,8))
    # mask = cv2.dilate(src_hsv, kernel)
    # kernel = cv2.getStructuringElement(cv2.MORPH_ELLIPSE, (8,8))
    # mask = cv2.erode(mask, kernel)

    contours, hierarchy = cv2.findContours(
        src_hsv, cv2.RETR_LIST, cv2.CHAIN_APPROX_NONE)
    for cnt in contours:
        area = cv2.contourArea(cnt)
        if area > 20000 and area < 40000:
            color_arr.append(color)

            cv2.imshow(color, src_hsv)


def get_circle(src):
    gray_image = cv2.cvtColor(src, cv2.COLOR_BGR2GRAY)
    circles = cv2.HoughCircles(gray_image, cv2.HOUGH_GRADIENT, 1, 50,
                               param1=150, param2=40, minRadius=80, maxRadius=150)  # 수치 변경 가능

    if circles is None:
        print("Not Circle")
        return 1

    if circles is not None:
        circles = np.uint16(np.around(circles))
        for i in circles[0, :]:
            cv2.circle(gray_image, (i[0], i[1]), i[2], (0, 255, 0), 2)

        cv2.imshow('detected circles', gray_image)

        # 원 감지됐을 때
        return get_line(src)    # return: 0 --> 원, 1 --> 원 아님


def get_line(src):
    imgray = cv2.cvtColor(src, cv2.COLOR_BGR2GRAY)
    edges = cv2.Canny(imgray, 50, 200)
    lines = cv2.HoughLinesP(edges, 1, np.pi/180, 10, None, 30, 2)

    if lines is None:
        print("Circle")
        return 0

    if lines is not None:
        for line in lines:
            x1, y1, x2, y2 = line[0]
            cv2.line(src, (x1, y1), (x2, y2), (0, 255, 0), 1)

        cv2.imshow('hough line', src)
        print("Not Circle")
        return 1


def categoryList(col):

    temp_list = [r[col].value for r in ws]
    del temp_list[0]
    temp_set = set(temp_list)
    name_list = list(temp_set)
    
    return name_list


def categoryData(col, name):
    total_list = []
    for r in ws.rows:
        temp_list = []
        cell_num = len(r)

        if col == 1:
            r_value = changeShape(r[col].value)

        elif col == 2:
            r_value = changeColor(r[col].value)

        if r_value == name:
            for n in range(0, cell_num):
                temp_list.append(r[n].value)
            if temp_list != []:
                total_list.append(temp_list)
    return total_list


def changeColor(name):
    if name == '분홍' or name == '노랑' or name == '주황' or name == '갈색' or name == '빨강':
        return "Red"
    if name == '연두' or name == '초록':
        return "Green"
    if name == '보라' or name == '파랑' or name == '자주':
        return "Blue"
    if name == '검정' or name == '회색':
        return "Black"
    if name == '하양':
        return "White"


def changeShape(name):
    if name == '원형':
        return "Circle"
    return "Not Circle"


def writeExcel(name, total_list):
    sht = wb.create_sheet(name)
    i = 1
    for data in total_list:
        data_length = len(data)
        for n in range(0, data_length):
            sht.cell(row=i, column=n+1).value = data[n]
        i = i+1


def findData(p_shape, p_color):
    shape = 1
    color = 2

    color_list = categoryData(color, p_color)
    
    shape_list = []
    for p in color_list:
        temp_list = []
        cell_num = len(p)
        p_value = changeShape(p[shape])

        if p_value == p_shape:
            for n in range(0, cell_num):
                temp_list.append(p[n])
            if temp_list != []:
                shape_list.append(temp_list)

    # 분류한 결과 return
    return shape_list
def getPillContour(src):
    mask = np.zeros(src.shape[:2], np.uint8)
    bgdModel = np.zeros((1, 65), np.float64)
    fgdModel = np.zeros((1, 65), np.float64)
    width = src.shape[1]
    height = src.shape[0]
    endx = width - (width // 10)
    endy = height - (height // 10)
    rect = (width//10, height//10, endx-width//10, endy-height//10)

    cv2.grabCut(src, mask, rect, bgdModel, fgdModel, 7, cv2.GC_INIT_WITH_RECT)
    mask2 = np.where((mask == 2) | (mask == 0), 0, 1).astype('uint8')
    src = src*mask2[:, :, np.newaxis]

    return src


def closing(src):
    kernel = np.ones((3, 3), np.uint8)
    result = cv2.dilate(src, kernel)
    result = cv2.erode(result, kernel)
    return result


def opening(src):
    kernel = np.ones((3, 3), np.uint8)
    result = cv2.erode(src, kernel)
    result = cv2.dilate(result, kernel)
    return result


def get_houghline(src, edges):
    lines = cv2.HoughLines(edges, 1, np.pi/180, 110)
    if lines is None:
        return src, [0]

    for i in range(len(lines)):
        for rho, theta in lines[i]:
            a = np.cos(theta)
            b = np.sin(theta)
            x0 = a*rho
            y0 = b*rho
            x1 = int(x0 + 1000*(-b))
            y1 = int(y0+1000*(a))
            x2 = int(x0 - 1000*(-b))
            y2 = int(y0 - 1000*(a))

            cv2.line(src, (x1, y1), (x2, y2), (0, 0, 255), 2)

    return src, lines


def dividerDetector(src, boundRect, lines):
    x, y, w, h = boundRect
    center = ((x + (x+w)) // 2, (y+(y+h)) // 2)


if __name__ == "__main__":
    src = cv2.imread('../195900032.jpg')
    if src is None:
        print('Image load failed!')
        sys.exit()

    # 색상
    color_arr = []
    get_color(src, redLow, redHigh, "Red", color_arr)
    get_color(src, greenLow, greenHigh, "Green", color_arr)
    get_color(src, blueLow, blueHigh, "Blue", color_arr)
    get_color(src, blackLow, blackHigh, "Black", color_arr)
    get_color(src, whiteLow, whiteHigh, "White", color_arr)

    # 모양
    circle = get_circle(src)

    # 색상, 모양에 따라 분류
    path = r"../pill_db.xlsx"
    wb = op.load_workbook(path)
    ws = wb.active

    findData(circle, color_arr[0])  # 분류한 결과 return
    wb.save("../분류결과파일.xlsx")  # 결과파일 저장

    print(color_arr)

    # 구분선
    image = cv2.GaussianBlur(src, (7, 7), 0)

    foreground = getPillContour(image)
    cv2.imshow("foreground", foreground)
    gray_foreground = cv2.cvtColor(foreground, cv2.COLOR_BGR2GRAY)
    clahe = cv2.createCLAHE(clipLimit=4.0, tileGridSize=(8, 8))
    clahe_foreground = clahe.apply(gray_foreground)
    ret, img_threshold = cv2.threshold(
        clahe_foreground, 100, 255, cv2.THRESH_BINARY)

    closing_image = closing(img_threshold)
    boundR = cv2.boundingRect(closing_image)

    canny = cv2.Canny(gray_foreground, 40, 150)
    ret, thresh = cv2.threshold(canny, 100, 255, cv2.THRESH_BINARY)

    # thresh = closing(thresh)
    # thresh = opening(thresh)
    cv2.imshow("Trhe", thresh)

    cv2.imshow("canny", canny)

    hough, lines = get_houghline(src, thresh)
    print("[ Divider Detector ]", lines)
    dividerDetector(hough, boundR, lines)
    cv2.waitKey(0)